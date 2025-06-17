import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Alignment
from datetime import datetime
import io
import os
import re # Import regex module

# --- Page Configuration ---
st.set_page_config(layout="centered", page_title="Đồng bộ dữ liệu SSE")

# --- File Paths ---
LOGO_PATH = "Logo.png"
DATA_FILE_PATH = "Data.xlsx"

# --- Headers for the output file ---
headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng",
           "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế",
           "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm",
           "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế",
           "Nhóm Hàng", "Ghi chú", "Tiền thuế"]

# --- Application Expiration Check ---
expiration_date = datetime(2025, 6, 26)
current_date = datetime.now()

if current_date > expiration_date:
    st.error("Có lỗi khi chạy chương trình, vui lòng liên hệ tác giả để được hỗ trợ!")
    st.info("Nguyễn Trọng Hoàn - 0902069469")
    st.stop()

# --- Helper Functions ---
def to_float(value):
    """Safely convert a value to a float."""
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def clean_string(s):
    """Remove extra whitespace from a string."""
    if pd.isna(s) or s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip()

# --- Static Data Loading ---
@st.cache_data
def get_static_data_from_excel(file_path):
    """Load static data and lookup tables from Data.xlsx."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        listbox_data = [clean_string(cell.value) for cell in ws['K'][3:] if cell.value]
        
        chxd_detail_map = {}
        store_specific_x_lookup = {}
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            chxd_name = clean_string(row[10])
            if chxd_name:
                chxd_detail_map[chxd_name] = {
                    'g5_val': row[15], 'h5_val': clean_string(row[17]).lower(),
                    'f5_val_full': clean_string(row[16]), 'b5_val': chxd_name
                }
                store_specific_x_lookup[chxd_name] = {
                    "xăng e5 ron 92-ii": row[11], "xăng ron 95-iii": row[12],
                    "dầu do 0,05s-ii": row[13], "dầu do 0,001s-v": row[14]
                }
        
        def create_lookup(min_r, max_r, min_c=9, max_c=10):
            return {clean_string(row[0]).lower(): row[1] for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True) if row[0] and row[1]}

        lookup_table = create_lookup(4, 7)
        tmt_lookup_table = {k: to_float(v) for k, v in create_lookup(10, 13).items()}
        s_lookup_table = create_lookup(29, 31)
        t_lookup_regular = create_lookup(33, 35)
        t_lookup_tmt = create_lookup(48, 50)
        v_lookup_table = create_lookup(53, 55)
        u_value = ws['J36'].value
        
        wb.close()
        
        return {
            "listbox_data": listbox_data, "lookup_table": lookup_table, "tmt_lookup_table": tmt_lookup_table,
            "s_lookup_table": s_lookup_table, "t_lookup_regular": t_lookup_regular, "t_lookup_tmt": t_lookup_tmt,
            "v_lookup_table": v_lookup_table, "u_value": u_value, "chxd_detail_map": chxd_detail_map,
            "store_specific_x_lookup": store_specific_x_lookup
        }
    except Exception as e:
        st.error(f"Lỗi nghiêm trọng khi đọc file cấu hình Data.xlsx: {e}")
        st.stop()

# --- Logic Functions ---
def add_tmt_summary_row(product_name_full, g5_val, s_lookup, t_lookup_tmt, v_lookup, u_val, h5_val, 
                        representative_date, representative_symbol, total_quantity_for_tmt, tmt_unit_value_for_summary, b5_val, customer_name_for_summary_row, x_lookup_for_store):
    new_tmt_row = [''] * len(headers)
    new_tmt_row[0], new_tmt_row[1], new_tmt_row[2] = g5_val, customer_name_for_summary_row, representative_date
    value_C, value_E = clean_string(representative_date), clean_string(representative_symbol)
    suffix_d = {"xăng e5 ron 92-ii": "1", "xăng ron 95-iii": "2", "dầu do 0,05s-ii": "3", "dầu do 0,001s-v": "4"}.get(product_name_full.lower(), "")
    if b5_val == "Nguyễn Huệ": new_tmt_row[3] = f"HNBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    elif b5_val == "Mai Linh": new_tmt_row[3] = f"MMBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    else: new_tmt_row[3] = f"{value_E[-2:]}BK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    new_tmt_row[4] = representative_symbol
    new_tmt_row[6], new_tmt_row[7], new_tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    new_tmt_row[9], new_tmt_row[12] = g5_val, total_quantity_for_tmt
    new_tmt_row[13] = tmt_unit_value_for_summary
    new_tmt_row[14] = round(to_float(total_quantity_for_tmt) * to_float(tmt_unit_value_for_summary), 0)
    new_tmt_row[17] = 10
    new_tmt_row[18] = s_lookup.get(h5_val, '')
    new_tmt_row[19] = t_lookup_tmt.get(h5_val, '')
    new_tmt_row[20], new_tmt_row[21] = u_val, v_lookup.get(h5_val, '')
    new_tmt_row[23] = x_lookup_for_store.get(product_name_full.lower(), '')
    new_tmt_row[31] = ""
    new_tmt_row[36] = round(to_float(total_quantity_for_tmt) * to_float(tmt_unit_value_for_summary) * 0.1, 0)
    for idx in [5,10,11,15,16,22,24,25,26,27,28,29,30,32,33,34,35]:
        if idx != 23 and idx < len(new_tmt_row): new_tmt_row[idx] = ''
    return new_tmt_row

def add_summary_row_for_no_invoice(data_for_summary_product, bkhd_source_ws_data, product_name, headers_list,
                    g5_val, b5_val, s_lookup, t_lookup, v_lookup, x_lookup_for_store, u_val, h5_val, common_lookup_table):
    new_row = [''] * len(headers_list)
    new_row[0], new_row[1] = g5_val, f"Khách hàng mua {product_name} không lấy hóa đơn"
    new_row[2] = data_for_summary_product[0][2] if data_for_summary_product else ""
    new_row[4] = data_for_summary_product[0][4] if data_for_summary_product else ""
    value_C, value_E = clean_string(new_row[2]), clean_string(new_row[4])
    suffix_d = {"xăng e5 ron 92-ii": "1", "xăng ron 95-iii": "2", "dầu do 0,05s-ii": "3", "dầu do 0,001s-v": "4"}.get(product_name.lower(), "")
    if b5_val == "Nguyễn Huệ": new_row[3] = f"HNBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    elif b5_val == "Mai Linh": new_row[3] = f"MMBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    else: new_row[3] = f"{value_E[-2:]}BK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    new_row[5] = f"Xuất bán lẻ theo hóa đơn số {new_row[3]}"
    new_row[6], new_row[7], new_row[8], new_row[9] = common_lookup_table.get(clean_string(product_name).lower(), ''), product_name, "Lít", g5_val
    total_M = sum(to_float(r[12]) for r in data_for_summary_product)
    new_row[12] = total_M
    new_row[13] = max((to_float(r[13]) for r in data_for_summary_product), default=0.0)
    tien_hang_hd = sum(to_float(r[11]) for r in bkhd_source_ws_data if len(r) > 8 and clean_string(r[5]) == "Người mua không lấy hóa đơn" and clean_string(r[8]) == product_name)
    price_per_liter = {"xăng e5 ron 92-ii": 1900, "xăng ron 95-iii": 2000, "dầu do 0,05s-ii": 1000, "dầu do 0,001s-v": 1000}.get(product_name.lower(), 0)
    new_row[14] = tien_hang_hd - round(total_M * price_per_liter, 0)
    new_row[17] = 10
    new_row[18], new_row[19] = s_lookup.get(h5_val, ''), t_lookup.get(h5_val, '')
    new_row[20], new_row[21] = u_val, v_lookup.get(h5_val, '')
    new_row[23] = x_lookup_for_store.get(clean_string(product_name).lower(), '')
    new_row[31] = f"Khách mua {product_name} không lấy hóa đơn"
    tienthue_hd_original = sum(to_float(row_bkhd[12]) for row_bkhd in bkhd_source_ws_data if len(row_bkhd) > 8 and clean_string(row_bkhd[5]) == "Người mua không lấy hóa đơn" and clean_string(row_bkhd[8]) == product_name)
    new_row[36] = tienthue_hd_original - round(total_M * price_per_liter * 0.1, 0) 
    return new_row

def create_per_invoice_tmt_row(original_row_data, tmt_value, g5_val, s_lookup, t_lookup_tmt, v_lookup, u_val, h5_val):
    tmt_row = list(original_row_data)
    tmt_row[6], tmt_row[7], tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    tmt_row[9] = g5_val
    tmt_row[13] = tmt_value
    tmt_row[14] = round(tmt_value * to_float(original_row_data[12]), 0)
    tmt_row[17] = 10
    tmt_row[18] = s_lookup.get(h5_val, '')
    tmt_row[19] = t_lookup_tmt.get(h5_val, '')
    tmt_row[20], tmt_row[21] = u_val, v_lookup.get(h5_val, '')
    tmt_row[31] = ""
    tmt_row[36] = round(tmt_value * to_float(original_row_data[12]) * 0.1, 0)
    for idx in [5, 10, 11, 15, 16, 22, 24, 25, 26, 27, 28, 29, 30, 32, 33, 34, 35]:
        if idx < len(tmt_row): tmt_row[idx] = ''
    return tmt_row

# --- Load Static Data ---
static_data = get_static_data_from_excel(DATA_FILE_PATH)
globals().update(static_data)

# --- UI Layout ---
col1, col2 = st.columns([1, 2])
with col1:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=180)
with col2:
    st.markdown("""
    <div style="display: flex; align-items: center; justify-content: center; height: 100px;">
        <h2 style="color: red; font-weight: bold; font-size: 24px; text-align: center; line-height: 1.1;">
            CÔNG TY CỔ PHẦN XĂNG DẦU<br>DẦU KHÍ NAM ĐỊNH
        </h2>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<h3 style='text-align: center; font-weight: bold;'>Công cụ đồng bộ dữ liệu lên phần mềm kế toán SSE</h3>", unsafe_allow_html=True)
st.markdown("""
<div class="blinking-warning" style="padding: 12px; background-color: #FFFACD; border: 1px solid #FFD700; border-radius: 8px; text-align: center;">
  <p style="color: #DC143C; font-weight: bold; margin: 0; font-size: 16px;">Lưu ý quan trọng: Nếu gặp lỗi, hãy mở file bảng kê và **Lưu (Save)** lại trước khi tải lên.</p>
</div><br>""", unsafe_allow_html=True)

selected_value = st.selectbox("Chọn CHXD:", options=[""] + listbox_data, key='selected_chxd')
uploaded_file = st.file_uploader("Tải lên file bảng kê hóa đơn (.xlsx)", type=["xlsx"])

st.markdown("---")
st.markdown("<p style='text-align: center; font-style: italic;'>Nếu gặp khó khăn trong quá trình sử dụng, hãy liên hệ: Nguyễn Trọng Hoàn - 0902069469</p>", unsafe_allow_html=True)

# --- Main Processing Logic ---
if st.button("Xử lý", key='process_button'):
    if not selected_value:
        st.warning("Vui lòng chọn một giá trị từ danh sách CHXD.")
    elif uploaded_file is None:
        st.warning("Vui lòng tải lên file bảng kê hóa đơn.")
    else:
        with st.spinner('Đang xử lý file, vui lòng chờ...'):
            try:
                # --- Step 1: Read File using openpyxl in read-only mode ---
                wb_in = load_workbook(uploaded_file, read_only=True, data_only=True)
                ws_in = wb_in.active
                
                all_rows_from_bkhd = [list(row) for row in ws_in.iter_rows(values_only=True)]
                bkhd_source_ws_data = all_rows_from_bkhd

                # --- Step 2: Process the Cleaned Data ---
                selected_value_normalized = clean_string(selected_value)
                chxd_details = chxd_detail_map.get(selected_value_normalized)
                
                if not chxd_details:
                    st.error(f"Không tìm thấy thông tin chi tiết cho CHXD: '{selected_value}'")
                    st.stop()
                
                g5_value, h5_value, f5_value_full, b5_value = chxd_details.get('g5_val'), chxd_details.get('h5_val'), chxd_details.get('f5_val_full'), chxd_details.get('b5_val')
                x_lookup_for_store = store_specific_x_lookup.get(selected_value_normalized, {})

                long_cells = [f"H{r_idx+1}" for r_idx, row in enumerate(all_rows_from_bkhd) if len(row) > 7 and row[7] and len(str(row[7])) > 128]
                if long_cells:
                    st.error("Địa chỉ trên ô " + ', '.join(long_cells) + " quá dài, hãy điều chỉnh và thử lại.")
                    st.stop()

                temp_bkhd_data = all_rows_from_bkhd[3:] if len(all_rows_from_bkhd) >= 4 else []
                vi_tri_cu_idx = [0, 1, 2, 3, 4, 5, 7, 6, 8, 10, 11, 13, 14, 16]
                intermediate_data = []
                for row in temp_bkhd_data:
                    if len(row) <= max(vi_tri_cu_idx): continue
                    new_row = [row[i] if i < len(row) else None for i in vi_tri_cu_idx]
                    date_val = new_row[3]
                    if isinstance(date_val, (datetime, pd.Timestamp)):
                        new_row[3] = date_val.strftime('%Y-%m-%d')
                    elif date_val:
                        try:
                            new_row[3] = pd.to_datetime(date_val, dayfirst=True).strftime('%Y-%m-%d')
                        except (ValueError, TypeError): pass
                    ma_kh = new_row[4]
                    new_row.append("No" if pd.isna(ma_kh) or len(clean_string(ma_kh)) > 9 else "Yes")
                    intermediate_data.append(new_row)

                if not intermediate_data:
                    st.error("Không có dữ liệu hợp lệ trong file bảng kê. File có thể trống hoặc sai định dạng.")
                    st.stop()

                b2_bkhd = clean_string(intermediate_data[0][1])
                f5_norm = clean_string(f5_value_full).lstrip('1')
                if f5_norm != b2_bkhd:
                    st.error(f"Bảng kê hóa đơn không phải của cửa hàng bạn chọn. Cửa hàng trong file: '{b2_bkhd}', Cửa hàng đã chọn: '{selected_value}'.")
                    st.stop()

                final_rows = [[''] * len(headers) for _ in range(4)] + [headers]
                all_tmt_rows = []
                no_invoice_rows = {p: [] for p in ["Xăng E5 RON 92-II", "Xăng RON 95-III", "Dầu DO 0,05S-II", "Dầu DO 0,001S-V"]}

                for row in intermediate_data:
                    upsse_row = [''] * len(headers)
                    upsse_row[0] = clean_string(row[4]) if row[-1] == 'Yes' and row[4] else g5_value
                    upsse_row[1], upsse_row[2] = clean_string(row[5]), row[3]
                    b_orig, c_orig = clean_string(row[1]), clean_string(row[2])
                    if b5_value == "Nguyễn Huệ": upsse_row[3] = f"HN{str(c_orig)[-6:]}"
                    elif b5_value == "Mai Linh": upsse_row[3] = f"MM{str(c_orig)[-6:]}"
                    else: upsse_row[3] = f"{str(b_orig)[-2:]}{str(c_orig)[-6:]}"
                    upsse_row[4] = f"1{b_orig}" if b_orig else ''
                    upsse_row[5] = f"Xuất bán lẻ theo hóa đơn số {upsse_row[3]}"
                    product_name = clean_string(row[8])
                    upsse_row[6] = lookup_table.get(product_name.lower(), '')
                    upsse_row[7] = product_name
                    upsse_row[8], upsse_row[9] = "Lít", g5_value
                    upsse_row[12] = to_float(row[9])
                    tmt_value = tmt_lookup_table.get(product_name.lower(), 0.0)
                    upsse_row[13] = round(to_float(row[10]) / 1.1 - tmt_value, 2)
                    upsse_row[14] = to_float(row[11]) - (tmt_value * upsse_row[12]) if upsse_row[12] else 0
                    upsse_row[17] = 10
                    upsse_row[18] = s_lookup_table.get(h5_value, '')
                    upsse_row[19] = t_lookup_regular.get(h5_value, '')
                    upsse_row[20], upsse_row[21] = u_value, v_lookup_table.get(h5_value, '')
                    upsse_row[23] = x_lookup_for_store.get(product_name.lower(), '')
                    upsse_row[31] = upsse_row[1]
                    upsse_row[32], upsse_row[33] = row[6], row[7]
                    upsse_row[36] = to_float(row[12]) - (upsse_row[12] * tmt_value * 0.1) if upsse_row[12] else 0

                    if upsse_row[1] == "Người mua không lấy hóa đơn" and product_name in no_invoice_rows:
                        no_invoice_rows[product_name].append(upsse_row)
                    else:
                        final_rows.append(upsse_row)
                        if tmt_value > 0 and upsse_row[12] > 0:
                            all_tmt_rows.append(create_per_invoice_tmt_row(upsse_row, tmt_value, g5_value, s_lookup_table, t_lookup_tmt, v_lookup_table, u_value, h5_value))

                for product_name, rows in no_invoice_rows.items():
                    if rows:
                        summary_row = add_summary_row_for_no_invoice(rows, bkhd_source_ws_data, product_name, headers, g5_value, b5_value, s_lookup_table, t_lookup_regular, v_lookup_table, x_lookup_for_store, u_value, h5_value, lookup_table)
                        final_rows.append(summary_row)
                        tmt_unit = tmt_lookup_table.get(product_name.lower(), 0)
                        total_qty = sum(to_float(r[12]) for r in rows)
                        if total_qty > 0:
                            all_tmt_rows.append(add_tmt_summary_row(product_name, g5_value, s_lookup_table, t_lookup_tmt, v_lookup_table, u_value, h5_value, summary_row[2], summary_row[4], total_qty, tmt_unit, b5_value, summary_row[1], x_lookup_for_store))

                final_rows.extend(all_tmt_rows)

                # --- Step 3: Create and Format Output File ---
                wb_out = Workbook()
                ws_out = wb_out.active
                for row_data in final_rows:
                    ws_out.append(row_data)

                text_style, date_style = NamedStyle(name="text_style", number_format='@'), NamedStyle(name="date_style", number_format='DD/MM/YYYY')
                exclude_cols = {3, 13, 14, 15, 18, 19, 20, 21, 22, 37}
                
                for r in range(1, ws_out.max_row + 1):
                    for c in range(1, ws_out.max_column + 1):
                        cell = ws_out.cell(row=r, column=c)
                        if not cell.value or clean_string(str(cell.value)) == "None": continue
                        if c == 3:
                            try:
                                cell.value = datetime.strptime(clean_string(str(cell.value)).split(" ")[0], '%Y-%m-%d').date()
                                cell.style = date_style
                            except (ValueError, TypeError): pass
                        elif c not in exclude_cols:
                            cell.style = text_style
                
                for c_idx in range(18, 23):
                     for r_idx in range(1, ws_out.max_row + 1):
                        ws_out.cell(row=r_idx, column=c_idx).number_format = '@'

                ws_out.column_dimensions['B'].width = 35
                ws_out.column_dimensions['C'].width = 12
                ws_out.column_dimensions['D'].width = 12

                output = io.BytesIO()
                wb_out.save(output)

                st.success("Tạo file UpSSE.xlsx thành công!")
                st.download_button(
                    label="Tải xuống file UpSSE.xlsx",
                    data=output.getvalue(),
                    file_name="UpSSE.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # --- THAY ĐỔI: Cập nhật khối xử lý lỗi ---
            except Exception as e:
                # Ghi lại lỗi kỹ thuật vào console để gỡ lỗi (nếu cần)
                print(f"Lỗi chi tiết: {e}")
                
                # Hiển thị thông báo lỗi thân thiện với người dùng
                st.error(
                    "**Lỗi xử lý file!**\n\n"
                    "Bạn cần kiểm tra lại file bảng kê. Hãy chắc chắn file được mở lên bằng Excel và **Lưu (Save)** lại trước khi đưa vào ứng dụng để xử lý.",
                    icon="🚨"
                )
                # Tùy chọn: hiển thị chi tiết lỗi kỹ thuật cho người dùng nâng cao
                # st.exception(e)

